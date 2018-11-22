import openpyxl

wb = openpyxl.load_workbook('candidates.xlsx')
sheet = wb['Candidates']
emails = [None] * (sheet.max_row - 1)
index = 0

for row in range(2, sheet.max_row + 1):
    emails[index] = sheet['B' + str(row)].value
    index += 1

tf = open("emails.txt", "w")

for index in range(0, len(emails)):
    tf.write(emails[index])
    tf.write(',')
    tf.write('\n')
